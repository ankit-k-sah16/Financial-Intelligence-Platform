"""
Peer Comparison Excel Exporter
------------------------------

Creates:

output/
    peer_comparison.xlsx

Features
--------
✓ One sheet per peer group
✓ 20 KPI columns
✓ Percentile Rank columns
✓ Conditional Formatting
✓ Benchmark Highlight
✓ Median Summary Row
"""

from pathlib import Path

import pandas as pd

from sqlalchemy import create_engine

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from config.setting import DB_PATH


class PeerComparisonExporter:

    OUTPUT = Path(
        "output/peer_comparison.xlsx"
    )

    GREEN = PatternFill(
        fill_type="solid",
        start_color="C6EFCE"
    )

    YELLOW = PatternFill(
        fill_type="solid",
        start_color="FFF2CC"
    )

    RED = PatternFill(
        fill_type="solid",
        start_color="F4CCCC"
    )

    GOLD = PatternFill(
        fill_type="solid",
        start_color="FFD966"
    )

    HEADER = PatternFill(
        fill_type="solid",
        start_color="4F81BD"
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF"
    )

    ALIGN = Alignment(
        horizontal="center"
    )

    def __init__(self):

        self.engine = create_engine(
            f"sqlite:///{DB_PATH}"
        )

    # -------------------------------------------------------
    # Load
    # -------------------------------------------------------

    def load(self):

        ratios = pd.read_sql(
            "SELECT * FROM stg_financial_ratios",
            self.engine
        )

        peers = pd.read_sql(
            "SELECT * FROM stg_peer_groups",
            self.engine
        )

        percentiles = pd.read_sql(
            "SELECT * FROM peer_percentiles",
            self.engine
        )

        self.df = ratios.merge(
            peers,
            on="company_id",
            how="left"
        )

        self.percentiles = percentiles

    # -------------------------------------------------------
    # Pivot Percentiles
    # -------------------------------------------------------

    def build_sheet(self, group):

        data = self.df[
            self.df["peer_group_name"] == group
        ].copy()

        pct = self.percentiles[
            self.percentiles["peer_group_name"] == group
        ]

        pivot = pct.pivot_table(

            index=["company_id", "year"],

            columns="metric",

            values="percentile_rank"

        )

        pivot.columns = [ f"{c}_Percentile" for c in pivot.columns  ]

        pivot = pivot.reset_index()

        data = data.merge(pivot, on=["company_id", "year"], how="left" )

        return data

    # -------------------------------------------------------
    # Export
    # -------------------------------------------------------

    def export(self):

        self.load()
        self.OUTPUT.parent.mkdir(exist_ok=True  )

        groups = sorted(self.df["peer_group_name"].dropna().unique())
        

        with pd.ExcelWriter( self.OUTPUT,engine="openpyxl") as writer:

            for group in groups:

                sheet = self.build_sheet(group)
                sheet.to_excel(writer, sheet_name=group[:31],index=False)

        wb = load_workbook(
            self.OUTPUT
        )

        for group in groups:
            ws = wb[group[:31]]

            self.style_sheet(ws)

            self.apply_percentile_colors(ws)

            self.highlight_benchmark(ws)

            self.add_summary_row(ws)

            self.auto_width(ws)

        wb.save(self.OUTPUT)

        print()
        print("=" * 60)
        print("Peer Comparison Report Generated")
        print(self.OUTPUT)
        print("=" * 60)

    # -------------------------------------------------------
    # Styling
    # -------------------------------------------------------

    def style_sheet(self, ws):

        for cell in ws[1]:

            cell.fill = self.HEADER

            cell.font = self.HEADER_FONT

            cell.alignment = self.ALIGN

        ws.freeze_panes = "A2"

        ws.auto_filter.ref = ws.dimensions

    # -------------------------------------------------------
    # Conditional Formatting
    # -------------------------------------------------------
    def apply_percentile_colors(self, ws):

        headers = { cell.value: cell.column for cell in ws[1]}

        percentile_cols = [ c for c in headers if "Percentile" in str(c)]

        for col_name in percentile_cols:

            col = headers[col_name]

            for row in range(2, ws.max_row):

                cell = ws.cell(row, col)

                if cell.value is None:

                    continue

                value = float(cell.value)

                if value >= 0.75:

                    cell.fill = self.GREEN

                elif value <= 0.25:

                    cell.fill = self.RED

                else:

                    cell.fill = self.YELLOW

    # -------------------------------------------------------
    # Benchmark Highlight
    # -------------------------------------------------------
    def highlight_benchmark(self, ws):

        headers = {cell.value: cell.column for cell in ws[1]   }

        if "is_benchmark" not in headers:

            return

        col = headers["is_benchmark"]

        for row in range(2, ws.max_row):

            if ws.cell(row, col).value:

                for c in range(1, ws.max_column + 1):

                    ws.cell(row, c).fill = self.GOLD

    # -------------------------------------------------------
    # Median Row
    # -------------------------------------------------------

    def add_summary_row(self, ws):

        last = ws.max_row + 1
        ws.cell(last, 1).value = "Peer Median"

        for col in range(2, ws.max_column + 1):

            values = []

            for row in range(2, ws.max_row):

                value = ws.cell(row, col).value

                if isinstance(value, (int, float)):

                    values.append(value)

            if values:
                ws.cell(last, col).value = round(pd.Series(values).median(),2)

    # -------------------------------------------------------
    # Width
    # -------------------------------------------------------
    def auto_width(self, ws):

        for column in ws.columns:
            width = max(len(str(cell.value))
                
                if cell.value is not None

                else 0

                for cell in column
            )

            ws.column_dimensions[column[0].column_letter].width = min(width + 3, 30 )

                

            

                

               

           