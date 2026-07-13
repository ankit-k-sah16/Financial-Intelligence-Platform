"""
Screener Excel Exporter
-----------------------

Creates:
output/
    screener_output.xlsx

Features
--------
✓ One worksheet per preset
✓ Top companies sorted by Composite Score
✓ Green cells = passes preset rule
✓ Red cells = fails preset rule
✓ Auto column width
✓ Freeze header
✓ Filters enabled
"""

from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment


class ScreenerExporter:

    OUTPUT_FILE = Path("output/screener_output.xlsx")

    GREEN = PatternFill(fill_type="solid",
       start_color="C6EFCE",
        end_color="C6EFCE",
    )

    RED = PatternFill(fill_type="solid",       
        start_color="FFC7CE",
        end_color="FFC7CE",
    )

    HEADER = PatternFill(fill_type="solid",
        start_color="4F81BD",
        end_color="4F81BD",
    )

    HEADER_FONT = Font(  bold=True,
        color="FFFFFF"
    )

    CENTER = Alignment( horizontal="center")
    
    # -------------------------------------------------------
    # Export
    # -------------------------------------------------------
    @classmethod
    def export(cls,preset_results,preset_filters,):

        """
        Parameters
        ----------
        preset_results
            Dictionary
            { preset_name : dataframe }
           
        preset_filters
            Dictionary
            { preset_name : filter_dictionary }
        """

        cls.OUTPUT_FILE.parent.mkdir(
            exist_ok=True
        )

        with pd.ExcelWriter( cls.OUTPUT_FILE, engine="openpyxl") as writer:

            for preset, df in preset_results.items():

                export_df = df.copy()

                export_df = export_df.sort_values("composite_quality_score",ascending=False,)
                    
                export_df.to_excel(writer,sheet_name=preset[:31],index=False,)

        workbook = load_workbook(cls.OUTPUT_FILE )
            
        for preset, filters in preset_filters.items():

            sheet = workbook[preset[:31]]

            cls.format_sheet(sheet)

            cls.apply_threshold_colors( sheet,filters,)
                           
            cls.auto_width(sheet)

        workbook.save(cls.OUTPUT_FILE)

        print()
        print("=" * 60)
        print("Screener Export Completed")
        print(cls.OUTPUT_FILE)
        print("=" * 60)

    # -------------------------------------------------------
    # Formatting
    # -------------------------------------------------------
    @classmethod
    def format_sheet(cls, sheet):

        for cell in sheet[1]:

            cell.fill = cls.HEADER

            cell.font = cls.HEADER_FONT

            cell.alignment = cls.CENTER

        sheet.freeze_panes = "A2"

        sheet.auto_filter.ref = sheet.dimensions

    # -------------------------------------------------------
    # Auto Width
    # -------------------------------------------------------
    @classmethod
    def auto_width(cls, sheet):

        for column in sheet.columns:

            length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:

                    length = max( length,len(str(cell.value)) )

                except Exception:
                    pass

            sheet.column_dimensions[column_letter].width = min(length + 3, 30)                 

    # -------------------------------------------------------
    # Conditional Formatting
    # -------------------------------------------------------
    @classmethod
    def apply_threshold_colors(  cls, sheet, filters,):

        header = {}

        for cell in sheet[1]:

            header[cell.value] = cell.column

        for metric, threshold in filters.items():

            column_name = cls.metric_to_column(metric )
  
            if column_name not in header:

                continue

            col = header[column_name]

            for row in range(2,sheet.max_row + 1, ):

                cell = sheet.cell(row=row,column=col,)

                value = cell.value

                if value is None:

                    continue
                
                try:

                    value = float(value)

                except Exception:

                    continue

                if metric.endswith("_min"):

                    if value >= threshold:

                        cell.fill = cls.GREEN

                    else:

                        cell.fill = cls.RED

                elif metric.endswith("_max"):

                    if value <= threshold:

                        cell.fill = cls.GREEN

                    else:

                        cell.fill = cls.RED

    # -------------------------------------------------------
    # Metric Mapping
    # -------------------------------------------------------
    @staticmethod
    def metric_to_column(metric):

        mapping = {

            "roe_min": "return_on_equity_pct",               
            "debt_to_equity_max": "debt_to_equity",               
            "fcf_min":"free_cash_flow_cr",            
            "revenue_cagr_5y_min": "revenue_cagr_5yr",
            "pat_cagr_5y_min": "pat_cagr_5yr",               
            "opm_min": "operating_profit_margin_pct",               
            "pe_max": "pe",
            "pb_max": "pb",            
            "dividend_yield_min":"dividend_yield",              
            "interest_coverage_ratio_min": "interest_coverage",            
            "market_cap_min":"market_cap",            
            "net_profit_min":"net_profit",
            "eps_cagr_min":"eps_cagr_5yr",            
            "asset_turnover_min":"asset_turnover",                
            "sales_min":"sales",               
            "dividend_payout_max": "dividend_payout_ratio_pct"
        }

        return mapping.get(metric)